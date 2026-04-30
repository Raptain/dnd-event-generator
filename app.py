from flask import Flask, render_template, jsonify
from openpyxl import load_workbook
import random
import os

app = Flask(__name__)

def load_events_from_excel():
    """Загружает события из Excel файла напрямую через openpyxl"""
    events = []
    try:
        # Получаем путь к файлу
        file_path = os.path.join(os.path.dirname(__file__), 'BaseDNDDesert.xlsx')
        
        # Загружаем Excel файл
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Получаем заголовки (первая строка)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        # Читаем данные со 2 строки
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            event = {}
            for i, header in enumerate(headers):
                if i < len(row) and row[i] is not None:
                    event[header] = row[i]
                else:
                    event[header] = ''
            
            if event.get('Название') and event.get('Название') != '':
                events.append(event)
        
        return events
    except Exception as e:
        print(f"Ошибка загрузки Excel: {e}")
        return []

# Загружаем события
EVENTS = load_events_from_excel()
print(f"✅ Загружено {len(EVENTS)} событий из Excel")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/random-event')
def random_event():
    if not EVENTS:
        return jsonify({'error': 'Нет событий в базе данных'}), 404
    
    event = random.choice(EVENTS)
    
    response = {
        'title': str(event.get('Название', 'Без названия')),
        'category': str(event.get('Категория', 'Неизвестно')),
        'difficulty': str(event.get('Сложность', 'Неизвестно')),
        'danger': int(event.get('Опасность (1–10)', 3)) if str(event.get('Опасность (1–10)', '')).isdigit() else 3,
        'reward': str(event.get('Награда', 'Нет награды')),
        'story': str(event.get('Сюжет', 'Описание отсутствует')),
        'development': str(event.get('Дальнейшее развитие', '')),
        'tags': str(event.get('Теги', ''))
    }
    
    return jsonify(response)

@app.route('/api/events-count')
def events_count():
    return jsonify({'count': len(EVENTS)})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)